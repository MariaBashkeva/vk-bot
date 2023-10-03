import requests
from bs4 import BeautifulSoup
import openpyxl
import vk_api
from vk_api.bot_longpoll import VkBotLongPoll, VkBotEventType
import datetime
import time
from vk_api.utils import get_random_id
vk_token = 'vk1.a.beNknCvpWT_qk-c7HfAuRZx9SddCxs9NwBTcOUJGzdr3qwOsaLjMl8OWC7ZuunH3DHTNQ0rahypVGJBA-ju4pw0f_jecbMTKdsedbFb0qPhUs_9QIs_jlWB1D8u8iyKbowVe2duZ594nJBTMEsdR0-3BJSTIu0T8DhCjaP0W5DBl37ACLmZaRqtTF1IsiNbBoxCxjjKOirXcoJI8whbwsA'  # Замените на ваш токен
group_id = '222303131'

# # Авторизация в ВКонтакте
vk_session = vk_api.VkApi(token=vk_token)
# vk = vk_session.get_api()
#
# # Долгpoll для получения новых сообщений
# longpoll = VkLongPoll(vk_session)

longpoll = VkBotLongPoll(vk_session, group_id)  # Подставьте свой
vk = vk_session.get_api()
# ...
def parse_url():
    try:
        url = "https://cfuv.ru/raspisanie-fakultativov-fiziko-tekhnicheskogo-instituta"
        response = None
        max_retries = 5  # Максимальное количество попыток
        retries = 0

        while retries < max_retries:
            try:
                response = requests.get(url, timeout=60)
                response.raise_for_status()  # Проверка на успешный статус ответа
                break
            except requests.exceptions.HTTPError as http_err:
                print(f"HTTP error occurred: {http_err}")
                break  # Прерываем цикл при HTTP-ошибках, но вы можете изменить логику обработки
            except requests.exceptions.ConnectionError as conn_err:
                print(f"Connection error occurred: {conn_err}")
                retries += 1
                if retries < max_retries:
                    print("Retrying in 5 seconds...")
                    time.sleep(5)  # Ждем 5 секунд перед повторной попыткой
                else:
                    print("Max retries reached. Exiting...")
                    break
            except Exception as err:
                print(f"An error occurred: {err}")
                retries += 1
                if retries < max_retries:
                    print("Retrying in 5 seconds...")
                    time.sleep(5)
                else:
                    print("Max retries reached. Exiting...")
                    break

        if response is not None and response.status_code == 200:
            # Ваш код обработки успешного ответа
            # ...
            soup = BeautifulSoup(response.text, 'html.parser')
            pr = soup.find_all("div", {"class": "card-body"})
            # print(pr)
            j = 0
            for i in pr:
                a = i.find_all("a")
                j += 1
                if j == 4:
                    return a[0]["href"]
        else:
            return False
    except Exception as ex:
        print(f"An error occurred: {ex}")
        return False










# Функция для парсинга расписания
def parse_schedule():
    try:
        schedule_url=parse_url()
        if not schedule_url:
            return ["Расписания еще нет" for i in range(5)]
        response = requests.get(schedule_url)
        dat = datetime.date.today()
        ned = True
        if( dat <= datetime.date(2023, 9, 15) or datetime.date(2023, 9, 23) <= dat <= datetime.date(
            2023, 9, 30) or datetime.date(2023, 10, 7) <= dat <= datetime.date(2023, 10, 13) or
            datetime.date(2023, 9, 11) <= dat <= datetime.date( 2023, 9, 15) or datetime.date(2023, 9, 11) <= dat <= datetime.date(2023, 9, 15) or
            datetime.date(2023, 9,11) <= dat <= datetime.date(2023, 9, 15)):
                ned = True
        else:
            ned = False

        if response.status_code == 200:

            soup = BeautifulSoup(response.text, 'html.parser')

            # Ваш код для парсинга расписания
            # Пример: находим и извлекаем информацию о расписании
            link = soup.find("div", {"class":"directDownload"}).find("a")["href"]
            res = requests.get(link, timeout=60)

            if res.status_code == 200:
                with open('data.xlsx', 'wb') as f:
                    f.write(res.content)

                # Загрузка Excel файла и обработка
                workbook = openpyxl.load_workbook('data.xlsx')
                sheet_lis=workbook.sheetnames
                schedule_info = []


                for sheet_name in sheet_lis:
                    if '3' in sheet_name and 'Ф,' in sheet_name:
                        sheet=workbook[sheet_name]
                        p=[]
                        t=[]
                        inf=[]
                        d=[]
                        c=0
                        for row in sheet.iter_rows(values_only=True):

                            for i in range(len(row)):
                                if row[i]==None:
                                    continue

                                if ned and 1<i<6:
                                    # print(row[1])
                                    if i==2:
                                        if row[i+1]!=None:
                                            # p.append(True)
                                            p.append((row[i]))
                                            c+=1
                                        else:
                                            if c!=0:
                                                d.append(c)
                                                c=0
                                    if i==3:
                                        t.append(row[i])
                                    if i==4:
                                        inf.append((row[i]))

                                if not ned and 11<i<15:
                                    print(row[i], i)
                                    if i == 12:
                                        if row[i+1]!=None:
                                            # p.append(True)
                                            p.append((row[i]))
                                            c+=1
                                        else:
                                            if c!=0:
                                                d.append(c)
                                                c=0
                                    if i==13:
                                        t.append(row[i])
                                    if i==14:
                                        inf.append(row[i])
                                schedule_info.append((p,t,inf))


                schedule = []
                t = 1
                inf = 2
                for i in range(1,len(schedule_info[0][0])):
                    schedule.append(

                        f"Пара: {schedule_info[0][0][i]}\nТип:{schedule_info[0][1][t]}\nПредмет:{schedule_info[0][2][inf]}\nПреподаватель:{schedule_info[0][2][inf + 1]}\nАдрес:{schedule_info[0][2][inf + 2]}\n\n")

                    t += 1
                    inf += 3
                d[0]-=1
                ans=[]
                k=0
                for i in d:
                    cur=[]
                    for j in range(i):
                        cur.append(schedule[k])
                        k+=1
                    if len(cur)!=0:
                        ans.append(cur)
                return ans
            else:
                return "Ошибка при скачивании файла"
        else:
            return "Ошибка при получении расписания"
    except:
        return ["Расписания еще нет" for i in range(5)]


schedule = parse_schedule()

print(schedule)
# Обработка новых сообщений
for event in longpoll.listen():
    try:
        if event.type == VkBotEventType.MESSAGE_NEW and event.from_chat:
            print(event.object['message']['text'])

            if event.object['message']['text'].lower() == '! обновить':
                schedule = parse_schedule()
                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message="Расписание обновлено",
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! помощь':
                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message="Напишите:\n! [день недели] для получения расписания на определенный день недели\n"
                    "! [обновить] для обновления расписания"
                            "! [сегодня] для получения расписания на сегодня\n"
                            "! [завтра] для получения расписания на завтра",
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! понедельник':

                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message=''.join(schedule[0]),
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! вторник':

                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message=''.join(schedule[1]),
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! среда':

                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message=''.join(schedule[2]),
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! четверг':

                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message=''.join(schedule[3]),
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! пятница':

                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message=''.join(schedule[4]) if len(schedule)>=5 else "Этот день свободный",
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! сегодня':
                current_date = datetime.date.today()

                # Получаем день недели (понедельник - 0, воскресенье - 6)
                day_of_week = current_date.weekday()

                # Преобразуем числовой день недели в текстовый формат

                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message=''.join(schedule[day_of_week]) if 0<=day_of_week<len(schedule) else "сегодня выходной",
                    random_id=get_random_id()
                )
            elif event.object['message']['text'].lower() == '! завтра':
                current_date = datetime.date.today()

                # Получаем день недели (понедельник - 0, воскресенье - 6)
                day_of_week = current_date.weekday()

                # Преобразуем числовой день недели в текстовый формат

                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message=''.join(schedule[day_of_week+1]) if 0<=day_of_week+1<len(schedule) else "завтра выходной",
                    random_id=get_random_id()

               )
            elif "!" in event.object['message']['text'].lower():
                vk.messages.send(
                    peer_id=event.object['message']['peer_id'],
                    message="такой команды нет, используйте ! [помощь]",
                    random_id=get_random_id()
                )
    except:
        continue

